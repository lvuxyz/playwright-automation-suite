"""
Đọc / ghi danh sách tài khoản từ file CSV, Excel hoặc TXT.

Định dạng cột bắt buộc:
    username, old_password, new_password

Cột tuỳ chọn (được giữ nguyên khi lưu lại):
    status      — kết quả lần chạy gần nhất (ok / fail / -)
    note        — ghi chú tuỳ ý

Định dạng TXT (mỗi dòng một tài khoản):
    username|old_password
    username|old_password|new_password
"""

from __future__ import annotations

import csv
import os
from dataclasses import dataclass, field, fields
from pathlib import Path
from typing import Iterable


# ── Kiểu dữ liệu một tài khoản ────────────────────────────────────────────

@dataclass
class Account:
    username:     str
    old_password: str
    new_password: str
    status:       str = "-"   # "-" | "ok" | "fail"
    note:         str = ""

    def as_row(self) -> list[str]:
        return [self.username, self.old_password, self.new_password,
                self.status, self.note]

    @classmethod
    def headers(cls) -> list[str]:
        return [f.name for f in fields(cls)]


# ── Đọc ───────────────────────────────────────────────────────────────────

def load_csv(path: str | Path) -> list[Account]:
    """Đọc danh sách tài khoản từ file CSV."""
    accounts: list[Account] = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                acc = Account(
                    username=row.get("username", "").strip(),
                    old_password=row.get("old_password", "").strip(),
                    new_password=row.get("new_password", "").strip(),
                    status=row.get("status", "-").strip() or "-",
                    note=row.get("note", "").strip(),
                )
                if acc.username:
                    accounts.append(acc)
            except Exception:
                continue
    return accounts


def load_excel(path: str | Path) -> list[Account]:
    """Đọc danh sách tài khoản từ file Excel (.xlsx)."""
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise ImportError("Cần cài openpyxl: pip install openpyxl") from exc

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    # Xác định cột dựa vào hàng tiêu đề (nếu có)
    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    col_map = {name: idx for idx, name in enumerate(header)}

    def _col(name: str, default: int) -> int:
        return col_map.get(name, default)

    start = 1 if any(k in col_map for k in ("username", "old_password")) else 0

    accounts: list[Account] = []
    for row in rows[start:]:
        if not row:
            continue
        def _get(idx: int) -> str:
            try:
                return str(row[idx]).strip() if row[idx] is not None else ""
            except IndexError:
                return ""

        acc = Account(
            username=_get(_col("username", 0)),
            old_password=_get(_col("old_password", 1)),
            new_password=_get(_col("new_password", 2)),
            status=_get(_col("status", 3)) or "-",
            note=_get(_col("note", 4)),
        )
        if acc.username:
            accounts.append(acc)
    return accounts


def load_txt(path: str | Path) -> list[Account]:
    """Đọc danh sách tài khoản từ file TXT.

    Mỗi dòng có định dạng:
        username|old_password
        username|old_password|new_password

    - Dòng trống hoặc chỉ có khoảng trắng sẽ bị bỏ qua.
    - Dòng không chứa dấu '|' sẽ bị bỏ qua.
    - Nếu password chứa ký tự '|', chỉ lần xuất hiện đầu tiên
      và thứ hai được dùng làm dấu phân cách (maxsplit=2).
    """
    accounts: list[Account] = []
    with open(path, encoding="utf-8-sig", errors="replace") as f:
        for raw in f:
            line = raw.strip()
            if not line or "|" not in line:
                continue

            parts = line.split("|", maxsplit=2)
            username     = parts[0].strip()
            old_password = parts[1].strip() if len(parts) > 1 else ""
            new_password = parts[2].strip() if len(parts) > 2 else ""

            if not username:
                continue

            accounts.append(Account(
                username=username,
                old_password=old_password,
                new_password=new_password,
            ))
    return accounts


def load_file(path: str | Path) -> list[Account]:
    """Tự động phát hiện định dạng file và tải."""
    ext = Path(path).suffix.lower()
    if ext in (".xlsx", ".xls"):
        return load_excel(path)
    if ext == ".txt":
        return load_txt(path)
    return load_csv(path)


# ── Ghi ───────────────────────────────────────────────────────────────────

def save_csv(accounts: Iterable[Account], path: str | Path) -> None:
    """Lưu danh sách tài khoản ra file CSV."""
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(Account.headers())
        for acc in accounts:
            writer.writerow(acc.as_row())


def save_excel(accounts: Iterable[Account], path: str | Path) -> None:
    """Lưu danh sách tài khoản ra file Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise ImportError("Cần cài openpyxl: pip install openpyxl") from exc

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Accounts"

    headers = Account.headers()
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="1565C0")
        cell.font = Font(bold=True, color="FFFFFF")

    for acc in accounts:
        ws.append(acc.as_row())

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            len(str(cell.value or "")) for cell in col
        ) + 4

    wb.save(path)


def save_file(accounts: Iterable[Account], path: str | Path) -> None:
    """Tự động chọn định dạng dựa theo đuôi file."""
    ext = Path(path).suffix.lower()
    if ext in (".xlsx", ".xls"):
        save_excel(accounts, path)
    else:
        save_csv(accounts, path)
